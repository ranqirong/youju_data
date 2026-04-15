#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
南宁新房合同数据处理工具 v2
基于新房合同导出接口的 5 个 sheet 生成每日和每周数据报表

Sheet 说明:
1. 新房合同 - 主表
2. 合同明细 - 详细合同信息
3. 解约和作废合同 - 退单数据
4. 应付费用项明细 - 业绩数据
5. 应收费用项明细 - 收款数据
"""

import json
import sys
import os
import importlib.util
from datetime import datetime, timedelta
from pathlib import Path
from collections import defaultdict

# 加载 erp_export 模块
script_dir = os.path.dirname(os.path.abspath(__file__))
erp_export_path = os.path.join(script_dir, 'erp-export.py')

spec = importlib.util.spec_from_file_location("erp_export", erp_export_path)
erp_export = importlib.util.module_from_spec(spec)
spec.loader.exec_module(erp_export)

ERPAPIClient = erp_export.ERPAPIClient
DEFAULT_CONFIG = erp_export.DEFAULT_CONFIG
get_user_credentials = erp_export.get_user_credentials


def load_excel_sheets(file_path):
    """加载 Excel 文件的 5 个 sheet（支持 inlineStr 表头）"""
    if not file_path or not os.path.exists(file_path):
        return None
    
    import zipfile
    import xml.etree.ElementTree as ET
    
    all_data = {}
    ns = {'ss': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
    
    try:
        with zipfile.ZipFile(file_path, 'r') as z:
            # 读取 workbook 获取 sheet 映射
            wb_xml = z.read('xl/workbook.xml')
            wb_root = ET.fromstring(wb_xml)
            
            sheets = wb_root.findall('.//ss:sheet', ns)
            sheet_map = {sheet.get('name'): f'xl/worksheets/sheet{i}.xml' 
                        for i, sheet in enumerate(sheets, 1)}
            
            # 读取每个 sheet
            for sheet_name in ['新房合同', '合同明细', '解约和作废合同', '应付费用项明细', '应收费用项明细']:
                sheet_file = sheet_map.get(sheet_name)
                if not sheet_file or sheet_file not in z.namelist():
                    print(f"  {sheet_name}: 未找到")
                    all_data[sheet_name] = []
                    continue
                
                # 读取 sheet XML
                xml_content = z.read(sheet_file)
                root = ET.fromstring(xml_content)
                
                rows = root.findall('.//ss:row', ns)
                
                # 读取表头（第一行，支持 inlineStr）
                headers = []
                if rows:
                    header_row = rows[0]
                    header_cells = header_row.findall('ss:c', ns)
                    for cell in header_cells:
                        # 尝试 inlineStr
                        t_elem = cell.find('ss:is/ss:t', ns)
                        if t_elem is not None and t_elem.text:
                            headers.append(t_elem.text)
                        else:
                            # 尝试普通 v
                            v_elem = cell.find('ss:v', ns)
                            headers.append(v_elem.text if v_elem is not None else '')
                
                # 读取数据行
                data_rows = []
                for row in rows[1:]:  # 跳过第一行（表头）
                    cells = row.findall('ss:c', ns)
                    values = []
                    for cell in cells:
                        val = ''
                        v_elem = cell.find('ss:v', ns)
                        if v_elem is not None and v_elem.text:
                            val = v_elem.text
                        else:
                            t_elem = cell.find('ss:is/ss:t', ns)
                            if t_elem is not None and t_elem.text:
                                val = t_elem.text
                        values.append(val)
                    if values and any(values):
                        # 转换为字典
                        record = {}
                        for i, val in enumerate(values):
                            if i < len(headers) and headers[i]:
                                record[headers[i]] = val
                        data_rows.append(record)
                
                print(f"  {sheet_name}: {len(data_rows)} 条记录")
                all_data[sheet_name] = data_rows
        
        return all_data
        
    except Exception as e:
        print(f"❌ 读取 Excel 失败：{e}")
        import traceback
        traceback.print_exc()
        return None
    
    # 方式 2: 使用 openpyxl 直接读取（处理 XML 问题）
    try:
        import openpyxl
        import zipfile
        import xml.etree.ElementTree as ET
        
        # 尝试修复 XML
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True, 
                                     keep_vba=False, keep_links=False)
        
        for sheet_name in wb.sheetnames:
            try:
                ws = wb[sheet_name]
                rows = []
                for row in ws.iter_rows(values_only=True):
                    if row and any(row):
                        rows.append(row)
                
                if len(rows) < 2:
                    print(f"  {sheet_name}: 无数据")
                    all_data[sheet_name] = []
                    continue
                
                headers = [str(h).strip() if h else '' for h in rows[0]]
                records = []
                for row in rows[1:]:
                    record = {}
                    for i, value in enumerate(row):
                        if i < len(headers) and headers[i]:
                            record[headers[i]] = value
                    records.append(record)
                
                print(f"  {sheet_name}: {len(records)} 条记录")
                all_data[sheet_name] = records
                
            except Exception as e:
                print(f"  ⚠️ {sheet_name}: {e}")
                all_data[sheet_name] = []
        
        return all_data
        
    except Exception as e:
        print(f"❌ 读取 Excel 失败：{e}")
        print("\n💡 Excel 文件可能有格式问题，尝试:")
        print("   1. 手动打开 Excel 并另存为新文件")
        print("   2. 或者使用其他工具转换格式")
        return None


def calculate_daily_data(all_data):
    """
    计算每日数据（基于列表格式的数据）
    
    假设合同明细 sheet 的列顺序:
    0:序号, 1:合同编号, 2:客户姓名, 3:身份证号, 4:联系电话,
    5:项目部层级2, 6:渠道部层级2, 7:授权渠道部层级2,
    8:项目部层级1, 9:渠道部层级1, 10:授权渠道部层级1,
    11:楼盘名称, 12:房号, 13:面积, 14:状态,
    15:认购金额, 16:基础应付金额, 17:签约时间, 18:备注
    """
    daily_data = []
    
    # 从"合同明细"sheet 获取主要数据（列表格式）
    contract_details = all_data.get('合同明细', [])
    
    for record in contract_details:
        # 根据列索引提取字段
        def get_val(idx):
            return record[idx] if idx < len(record) and record[idx] else ''
        
        project_dept_l2 = get_val(5)
        channel_dept_l2 = get_val(6)
        auth_channel_dept_l2 = get_val(7)
        project_dept_l1 = get_val(8)
        channel_dept_l1 = get_val(9)
        auth_channel_dept_l1 = get_val(10)
        building_name = get_val(11)
        
        # 获取状态和金额
        status = get_val(14)
        base_pay_amount = float(get_val(16) or 0)
        
        # 根据状态计算套数和业绩
        if status in ['认购', '已认购']:
            subscribe_units = 1
            subscribe_amount = base_pay_amount
            sign_units = 0
            sign_amount = 0
        elif status in ['签约', '已签约']:
            subscribe_units = 0
            subscribe_amount = 0
            sign_units = 1
            sign_amount = base_pay_amount
        else:
            subscribe_units = 0
            subscribe_amount = 0
            sign_units = 0
            sign_amount = 0
        
        # 获取签约时间（可能是时间戳或日期字符串）
        sign_tm_raw = get_val(17)
        sign_date = ''
        if sign_tm_raw:
            try:
                sign_tm = float(sign_tm_raw)
                sign_date = datetime.fromtimestamp(sign_tm / 1000).strftime('%Y-%m-%d')
            except:
                sign_date = str(sign_tm_raw)[:10]
        
        daily_data.append({
            '日期': sign_date,
            '项目部层级2': project_dept_l2,
            '渠道部层级2': channel_dept_l2,
            '授权渠道部层级2': auth_channel_dept_l2,
            '项目部层级1': project_dept_l1,
            '渠道部层级1': channel_dept_l1,
            '授权渠道部层级1': auth_channel_dept_l1,
            '楼盘名称': building_name,
            '认购套数': subscribe_units,
            '认购业绩': subscribe_amount,
            '签约套数': sign_units,
            '签约业绩': sign_amount,
            '状态': status
        })
    
    return daily_data


def calculate_weekly_data(all_data):
    """计算每周数据（基于列表格式）"""
    weekly_agg = defaultdict(lambda: {
        'subscribeUnits': 0,
        'subscribeAmount': 0,
        'signUnitsL1': 0,
        'signUnitsL2': 0,
        'signAmount': 0,
        'cancelUnits': 0,
        'cancelAmount': 0,
        'excludeUnits': 0,
        'excludeAmount': 0,
        'target': 0,
    })
    
    contract_details = all_data.get('合同明细', [])
    
    for record in contract_details:
        def get_val(idx):
            return record[idx] if idx < len(record) and record[idx] else ''
        
        # 获取签约时间（列 17）
        sign_tm_raw = get_val(17)
        week_key = '未知'
        if sign_tm_raw:
            try:
                sign_tm = float(sign_tm_raw)
                sign_date = datetime.fromtimestamp(sign_tm / 1000)
                week_start = sign_date - timedelta(days=sign_date.weekday())
                week_key = week_start.strftime('%Y-%m-%d')
            except:
                pass
        
        project_dept_l2 = get_val(5)
        project_dept_l1 = get_val(8)
        key = f"{week_key}|{project_dept_l2}|{project_dept_l1}"
        
        status = get_val(14)
        base_pay_amount = float(get_val(16) or 0)
        
        if status in ['认购', '已认购']:
            weekly_agg[key]['subscribeUnits'] += 1
            weekly_agg[key]['subscribeAmount'] += base_pay_amount
        elif status in ['签约', '已签约']:
            channel_dept_l1 = get_val(9)
            channel_dept_l2 = get_val(6)
            
            if channel_dept_l1:
                weekly_agg[key]['signUnitsL1'] += 1
            if channel_dept_l2:
                weekly_agg[key]['signUnitsL2'] += 1
            
            weekly_agg[key]['signAmount'] += base_pay_amount
        elif status in ['退单', '已退单', '解约', '作废']:
            weekly_agg[key]['cancelUnits'] += 1
            weekly_agg[key]['cancelAmount'] += base_pay_amount
        elif status in ['剔除', '已剔除']:
            weekly_agg[key]['excludeUnits'] += 1
            weekly_agg[key]['excludeAmount'] += base_pay_amount
    
    weekly_data = []
    for key, agg in weekly_agg.items():
        week, project_dept_l2, project_dept_l1 = key.split('|')
        
        total_units = agg['signUnitsL1'] + agg['signUnitsL2']
        target = agg['target'] if agg['target'] > 0 else 1
        completion_rate = (agg['signAmount'] / target * 100) if target > 0 else 0
        
        weekly_data.append({
            '周': week,
            '项目部层级2': project_dept_l2,
            '项目部层级1': project_dept_l1,
            '一级认购套数': agg['subscribeUnits'],
            '一级认购业绩': agg['subscribeAmount'],
            '一级签约套数': agg['signUnitsL1'],
            '二级签约套数': agg['signUnitsL2'],
            '套数合计': total_units,
            '一级业绩': agg['signAmount'],
            '退单套数': agg['cancelUnits'],
            '退单业绩': agg['cancelAmount'],
            '剔除退单套数': agg['excludeUnits'],
            '剔除退单业绩': agg['excludeAmount'],
            '目标': agg['target'],
            '完成率': f"{completion_rate:.2f}%"
        })
    
    return weekly_data


def export_to_excel(data, output_file, sheet_name='数据'):
    """导出数据到 Excel"""
    try:
        import pandas as pd
        
        if not data:
            # 创建空 DataFrame
            df = pd.DataFrame()
        else:
            df = pd.DataFrame(data)
        
        # 创建 Excel writer
        output_path = Path(output_file)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        # 如果文件存在，追加 sheet；否则创建新文件
        if output_path.exists():
            with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        else:
            df.to_excel(output_file, sheet_name=sheet_name, index=False)
        
        print(f"✅ Excel 导出成功：{output_file}")
        return True
    except ImportError as e:
        print(f"⚠️ 未安装 pandas 或 openpyxl: {e}")
        print("   请运行：pip3 install pandas openpyxl")
        
        # 保存为 JSON
        json_file = output_file.replace('.xlsx', '.json')
        with open(json_file, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        print(f"✅ JSON 导出成功：{json_file}")
        return False
    except Exception as e:
        print(f"❌ Excel 导出失败：{e}")
        return False


def main():
    print("="*70)
    print("南宁新房合同数据处理工具 v2")
    print("="*70)
    print(f"当前时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    # 【安全检查】获取用户凭证
    phone, password = get_user_credentials()
    
    if not phone or not password:
        print("\n" + "="*70)
        print("❌ 错误：未找到 ERP 登录凭证")
        print("="*70)
        return
    
    # 创建客户端
    print(f"\n使用账号：{phone}")
    client = ERPAPIClient(
        phone=phone,
        password=password
    )
    
    # 登录
    if not client.login():
        print("\n❌ 登录失败")
        return
    
    # 获取用户信息
    print("\n获取用户信息...")
    user_info = client.get_user_info()
    if not user_info:
        print("\n❌ 程序终止：用户信息获取失败")
        return
    
    print(f"✅ 用户：{user_info.get('userName')}")
    
    # 导出新房合同数据
    print("\n" + "="*70)
    print("步骤 1: 导出新房合同数据")
    print("="*70)
    
    # 支持命令行参数指定日期
    import sys
    query_date = datetime.now()  # 默认今日
    
    if len(sys.argv) > 1:
        try:
            # 支持格式：YYYY-MM-DD 或 YYYYMMDD
            date_str = sys.argv[1]
            if '-' in date_str:
                query_date = datetime.strptime(date_str, '%Y-%m-%d')
            else:
                query_date = datetime.strptime(date_str, '%Y%m%d')
            print(f"查询日期：{query_date.strftime('%Y-%m-%d')}")
        except:
            print(f"⚠️ 日期格式错误，使用今日：{query_date.strftime('%Y-%m-%d')}")
    else:
        print(f"查询日期：{query_date.strftime('%Y-%m-%d')} (今日)")
    
    start_time = int(query_date.replace(hour=0, minute=0, second=0, microsecond=0).timestamp() * 1000)
    end_time = int((query_date + timedelta(days=1)).timestamp() * 1000) - 1
    
    result = client.export_new_house_contracts(
        start_time=start_time,
        end_time=end_time,
        city_code='450100',  # 南宁
        organ_id='625864877560328320'
    )
    
    if not result or not result.get('success'):
        print("\n❌ 新房合同数据导出失败")
        return
    
    # 处理数据
    print("\n" + "="*70)
    print("步骤 2: 处理 Excel 数据")
    print("="*70)
    
    excel_file = result.get('excel_file')
    print(f"\n读取 Excel 文件：{excel_file}")
    all_data = load_excel_sheets(excel_file)
    
    if not all_data:
        print("\n❌ 无法读取 Excel 数据")
        return
    
    # 统计各 sheet 数据量
    total_records = sum(len(records) for records in all_data.values())
    print(f"\n✅ 共读取 {total_records} 条记录")
    
    # 计算每日数据
    print("\n" + "="*70)
    print("步骤 3: 计算每日数据")
    print("="*70)
    daily_data = calculate_daily_data(all_data)
    print(f"✅ 生成 {len(daily_data)} 条每日数据")
    
    # 计算每周数据
    print("\n" + "="*70)
    print("步骤 4: 计算每周数据")
    print("="*70)
    weekly_data = calculate_weekly_data(all_data)
    print(f"✅ 生成 {len(weekly_data)} 条每周数据")
    
    # 导出 Excel
    output_dir = Path('~/Desktop/ERP 导出/南宁新房').expanduser()
    output_dir.mkdir(parents=True, exist_ok=True)
    
    timestamp = datetime.now().strftime('%Y%m%d')
    daily_file = output_dir / f"南宁新房每日数据_{timestamp}.xlsx"
    weekly_file = output_dir / f"南宁新房每周数据_{timestamp}.xlsx"
    
    print(f"\n" + "="*70)
    print("步骤 5: 导出报表")
    print("="*70)
    
    print(f"\n导出每日数据到：{daily_file}")
    export_to_excel(daily_data, str(daily_file), '每日数据')
    
    print(f"导出每周数据到：{weekly_file}")
    export_to_excel(weekly_data, str(weekly_file), '每周数据')
    
    # 显示统计摘要
    print("\n" + "="*70)
    print("📊 数据统计摘要")
    print("="*70)
    
    if daily_data:
        total_subscribe = sum(d['认购套数'] for d in daily_data)
        total_subscribe_amount = sum(d['认购业绩'] for d in daily_data)
        total_sign = sum(d['签约套数'] for d in daily_data)
        total_sign_amount = sum(d['签约业绩'] for d in daily_data)
        
        print(f"\n今日总计:")
        print(f"   认购套数：{total_subscribe}")
        print(f"   认购业绩：¥{total_subscribe_amount:,.2f}")
        print(f"   签约套数：{total_sign}")
        print(f"   签约业绩：¥{total_sign_amount:,.2f}")
    
    if weekly_data:
        print(f"\n本周统计:")
        print(f"   数据维度：{len(weekly_data)} 条")
    
    print("\n" + "="*70)
    print("✅ 数据处理完成！")
    print("="*70)
    print(f"\n📁 输出文件:")
    print(f"   {daily_file}")
    print(f"   {weekly_file}")
    print("="*70)


if __name__ == "__main__":
    main()
