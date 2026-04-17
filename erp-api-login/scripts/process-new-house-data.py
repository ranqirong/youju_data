#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
南宁新房合同数据处理工具
基于新房合同导出接口数据生成每日和每周数据报表
"""

import json
import sys
import os
import importlib.util
from datetime import datetime, timedelta
from pathlib import Path
from collections import defaultdict

# 加载 erp_export 模块
script_dir = os.path.dirname(os.path.abspath(__file__))
erp_export_path = os.path.join(script_dir, 'erp-export.py')

spec = importlib.util.spec_from_file_location("erp_export", erp_export_path)
erp_export = importlib.util.module_from_spec(spec)
spec.loader.exec_module(erp_export)

ERPAPIClient = erp_export.ERPAPIClient
DEFAULT_CONFIG = erp_export.DEFAULT_CONFIG
get_user_credentials = erp_export.get_user_credentials


def load_contract_data(file_path=None):
    """加载新房合同数据（从 Excel 文件）"""
    if file_path and os.path.exists(file_path):
        try:
            import openpyxl
            # 使用 openpyxl 直接读取
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            
            # 读取所有 sheet 的数据
            all_records = []
            for sheet_name in wb.sheetnames:
                try:
                    ws = wb[sheet_name]
                    rows = list(ws.iter_rows(values_only=True))
                    if len(rows) < 2:
                        continue
                    
                    # 第一行是表头
                    headers = [str(h) if h else '' for h in rows[0]]
                    
                    # 转换为字典
                    for row in rows[1:]:
                        if row and any(row):  # 跳过空行
                            record = {}
                            for i, value in enumerate(row):
                                if i < len(headers):
                                    record[headers[i]] = value
                            record['_sheet'] = sheet_name  # 添加 sheet 标记
                            all_records.append(record)
                    
                    print(f"  {sheet_name}: {len(rows)-1} 条记录")
                except Exception as e:
                    print(f"  ⚠️ 读取 {sheet_name} 失败：{e}")
            
            print(f"✅ 读取 Excel 文件：共 {len(all_records)} 条记录")
            return all_records
            
        except ImportError:
            print("❌ 未安装 openpyxl")
            print("   请运行：pip3 install openpyxl")
            return None
        except Exception as e:
            print(f"❌ 读取 Excel 失败：{e}")
            return None
    return None


def calculate_daily_data(records):
    """
    计算每日数据
    
    包含:
    - 项目部层级2
    - 渠道部层级2
    - 授权渠道部层级2
    - 项目部层级1
    - 渠道部层级1
    - 授权渠道部层级1
    - 楼盘名称
    - 认购套数 (根据认购状态计算)
    - 认购业绩
    - 签约套数 (根据签约状态计算)
    - 签约业绩
    """
    daily_data = []
    
    for record in records:
        # 提取字段
        project_dept_l2 = record.get('projectDeptL2', '')  # 项目部层级2
        channel_dept_l2 = record.get('channelDeptL2', '')  # 渠道部层级2
        auth_channel_dept_l2 = record.get('authChannelDeptL2', '')  # 授权渠道部层级2
        project_dept_l1 = record.get('projectDeptL1', '')  # 项目部层级1
        channel_dept_l1 = record.get('channelDeptL1', '')  # 渠道部层级1
        auth_channel_dept_l1 = record.get('authChannelDeptL1', '')  # 授权渠道部层级1
        building_name = record.get('buildingName', '')  # 楼盘名称
        
        # 根据状态计算套数和业绩
        status = record.get('status', '')
        base_pay_amount = float(record.get('basePayAmount', 0) or 0)  # 基础应付金额（业绩）
        
        # 认购状态
        if status in ['认购', '已认购']:
            subscribe_units = 1
            subscribe_amount = base_pay_amount
            sign_units = 0
            sign_amount = 0
        # 签约状态
        elif status in ['签约', '已签约']:
            subscribe_units = 0
            subscribe_amount = 0
            sign_units = 1
            sign_amount = base_pay_amount
        else:
            subscribe_units = 0
            subscribe_amount = 0
            sign_units = 0
            sign_amount = 0
        
        daily_data.append({
            'date': record.get('signTm', '')[:10] if record.get('signTm') else '',
            'projectDeptL2': project_dept_l2,
            'channelDeptL2': channel_dept_l2,
            'authChannelDeptL2': auth_channel_dept_l2,
            'projectDeptL1': project_dept_l1,
            'channelDeptL1': channel_dept_l1,
            'authChannelDeptL1': auth_channel_dept_l1,
            'buildingName': building_name,
            'subscribeUnits': subscribe_units,
            'subscribeAmount': subscribe_amount,
            'signUnits': sign_units,
            'signAmount': sign_amount,
            'status': status
        })
    
    return daily_data


def calculate_weekly_data(records):
    """
    计算每周数据
    
    包含:
    - 项目部层级2
    - 项目部层级1
    - 一级认购套数（根据认购状态计算）
    - 一级认购业绩（基础应付金额）
    - 一级签约套数（根据签约状态计算）
    - 二级签约套数（根据签约状态计算）
    - 套数合计
    - 一级业绩（基础应付金额）
    - 退单套数
    - 退单业绩
    - 剔除退单套数
    - 剔除退单业绩
    - 目标
    - 完成率
    """
    # 按周和项目部聚合
    weekly_agg = defaultdict(lambda: {
        'subscribeUnits': 0,
        'subscribeAmount': 0,
        'signUnitsL1': 0,
        'signUnitsL2': 0,
        'signAmount': 0,
        'cancelUnits': 0,
        'cancelAmount': 0,
        'excludeUnits': 0,
        'excludeAmount': 0,
        'target': 0,
    })
    
    for record in records:
        # 获取周信息（以签约时间为准）
        sign_tm = record.get('signTm', '')
        if sign_tm:
            sign_date = datetime.fromtimestamp(int(sign_tm) / 1000)
            # 计算周开始时间（周一）
            week_start = sign_date - timedelta(days=sign_date.weekday())
            week_key = week_start.strftime('%Y-%m-%d')
        else:
            week_key = '未知'
        
        project_dept_l2 = record.get('projectDeptL2', '')
        project_dept_l1 = record.get('projectDeptL1', '')
        key = f"{week_key}|{project_dept_l2}|{project_dept_l1}"
        
        status = record.get('status', '')
        base_pay_amount = float(record.get('basePayAmount', 0) or 0)
        
        # 根据状态统计
        if status in ['认购', '已认购']:
            weekly_agg[key]['subscribeUnits'] += 1
            weekly_agg[key]['subscribeAmount'] += base_pay_amount
        elif status in ['签约', '已签约']:
            # 根据层级判断一级还是二级签约
            if record.get('channelDeptL1'):
                weekly_agg[key]['signUnitsL1'] += 1
            if record.get('channelDeptL2'):
                weekly_agg[key]['signUnitsL2'] += 1
            weekly_agg[key]['signAmount'] += base_pay_amount
        elif status in ['退单', '已退单']:
            weekly_agg[key]['cancelUnits'] += 1
            weekly_agg[key]['cancelAmount'] += base_pay_amount
        elif status in ['剔除', '已剔除']:
            weekly_agg[key]['excludeUnits'] += 1
            weekly_agg[key]['excludeAmount'] += base_pay_amount
    
    # 生成每周数据
    weekly_data = []
    for key, agg in weekly_agg.items():
        week, project_dept_l2, project_dept_l1 = key.split('|')
        
        total_units = agg['signUnitsL1'] + agg['signUnitsL2']
        target = agg['target'] if agg['target'] > 0 else 1  # 避免除零
        completion_rate = (agg['signAmount'] / target * 100) if target > 0 else 0
        
        weekly_data.append({
            'week': week,
            'projectDeptL2': project_dept_l2,
            'projectDeptL1': project_dept_l1,
            'subscribeUnits': agg['subscribeUnits'],
            'subscribeAmount': agg['subscribeAmount'],
            'signUnitsL1': agg['signUnitsL1'],
            'signUnitsL2': agg['signUnitsL2'],
            'totalUnits': total_units,
            'signAmount': agg['signAmount'],
            'cancelUnits': agg['cancelUnits'],
            'cancelAmount': agg['cancelAmount'],
            'excludeUnits': agg['excludeUnits'],
            'excludeAmount': agg['excludeAmount'],
            'target': agg['target'],
            'completionRate': f"{completion_rate:.2f}%"
        })
    
    return weekly_data


def export_to_excel(data, output_file, sheet_name='数据'):
    """导出数据到 Excel"""
    try:
        import pandas as pd
        
        df = pd.DataFrame(data)
        
        # 创建 Excel writer
        output_path = Path(output_file)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        # 如果文件存在，追加 sheet；否则创建新文件
        if output_path.exists():
            with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        else:
            df.to_excel(output_file, sheet_name=sheet_name, index=False)
        
        print(f"✅ Excel 导出成功：{output_file}")
        return True
    except ImportError:
        print("⚠️ 未安装 pandas 或 openpyxl，使用 JSON 格式保存")
        with open(output_file.replace('.xlsx', '.json'), 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        return False
    except Exception as e:
        print(f"❌ Excel 导出失败：{e}")
        return False


def main():
    print("="*70)
    print("南宁新房合同数据处理工具")
    print("="*70)
    print(f"当前时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    # 【安全检查】获取用户凭证
    phone, password = get_user_credentials()
    
    if not phone or not password:
        print("\n" + "="*70)
        print("❌ 错误：未找到 ERP 登录凭证")
        print("="*70)
        print("\n💡 请配置凭证后重试:")
        print("   1. 编辑：~/.openclaw/workspace/configs/erp-credentials.env")
        print("   2. 添加:")
        print("      ERP_USERNAME=你的手机号")
        print("      ERP_PASSWORD=你的密码")
        print("="*70)
        return
    
    # 创建客户端
    print(f"\n使用账号：{phone}")
    client = ERPAPIClient(
        phone=phone,
        password=password
    )
    
    # 登录
    if not client.login():
        print("\n❌ 登录失败")
        return
    
    # 获取用户信息
    print("\n获取用户信息...")
    user_info = client.get_user_info()
    if not user_info:
        print("\n❌ 程序终止：用户信息获取失败")
        return
    
    print(f"✅ 用户：{user_info.get('userName')}")
    
    # 导出今日新房合同数据
    print("\n" + "="*70)
    print("步骤 1: 导出今日新房合同数据")
    print("="*70)
    
    today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    start_time = int(today.timestamp() * 1000)
    end_time = int((today + timedelta(days=1)).timestamp() * 1000) - 1
    
    result = client.export_new_house_contracts(
        start_time=start_time,
        end_time=end_time,
        city_code='450100',  # 南宁
        organ_id='625864877560328320'
    )
    
    if not result or not result.get('success'):
        print("\n❌ 新房合同数据导出失败")
        return
    
    # 处理数据
    print("\n" + "="*70)
    print("步骤 2: 处理数据生成报表")
    print("="*70)
    
    # 从 Excel 文件读取数据
    excel_file = result.get('excel_file')
    records = load_contract_data(excel_file)
    
    if not records:
        print("\n⚠️ 今日没有新房合同数据")
        print("\n💡 可能原因:")
        print("   1. 今日确实没有新房合同")
        print("   2. 账号权限不足，无法查看该城市数据")
        print("   3. 查询条件不正确")
        
        # 生成空报表
        output_dir = Path('~/Desktop/ERP 导出/南宁新房').expanduser()
        output_dir.mkdir(parents=True, exist_ok=True)
        
        timestamp = datetime.now().strftime('%Y%m%d')
        daily_file = output_dir / f"南宁新房每日数据_{timestamp}.xlsx"
        weekly_file = output_dir / f"南宁新房每周数据_{timestamp}.xlsx"
        
        # 创建空 Excel 文件
        try:
            import pandas as pd
            empty_df = pd.DataFrame(columns=[
                '日期', '项目部层级2', '渠道部层级2', '授权渠道部层级2',
                '项目部层级1', '渠道部层级1', '授权渠道部层级1', '楼盘名称',
                '认购套数', '认购业绩', '签约套数', '签约业绩'
            ])
            empty_df.to_excel(daily_file, index=False)
            empty_df.to_excel(weekly_file, index=False)
            
            print(f"\n✅ 已生成空报表:")
            print(f"   {daily_file}")
            print(f"   {weekly_file}")
        except:
            pass
        
        return
    
    print(f"✅ 获取到 {len(records)} 条合同记录")
    
    # 计算每日数据
    print("\n计算每日数据...")
    daily_data = calculate_daily_data(records)
    
    # 计算每周数据
    print("计算每周数据...")
    weekly_data = calculate_weekly_data(records)
    
    # 导出 Excel
    output_dir = Path('~/Desktop/ERP 导出/南宁新房').expanduser()
    output_dir.mkdir(parents=True, exist_ok=True)
    
    timestamp = datetime.now().strftime('%Y%m%d')
    daily_file = output_dir / f"南宁新房每日数据_{timestamp}.xlsx"
    weekly_file = output_dir / f"南宁新房每周数据_{timestamp}.xlsx"
    
    print(f"\n导出每日数据到：{daily_file}")
    export_to_excel(daily_data, str(daily_file), '每日数据')
    
    print(f"导出每周数据到：{weekly_file}")
    export_to_excel(weekly_data, str(weekly_file), '每周数据')
    
    # 显示统计摘要
    print("\n" + "="*70)
    print("📊 数据统计摘要")
    print("="*70)
    
    total_subscribe = sum(d['subscribeUnits'] for d in daily_data)
    total_subscribe_amount = sum(d['subscribeAmount'] for d in daily_data)
    total_sign = sum(d['signUnits'] for d in daily_data)
    total_sign_amount = sum(d['signAmount'] for d in daily_data)
    
    print(f"\n今日总计:")
    print(f"   认购套数：{total_subscribe}")
    print(f"   认购业绩：¥{total_subscribe_amount:,.2f}")
    print(f"   签约套数：{total_sign}")
    print(f"   签约业绩：¥{total_sign_amount:,.2f}")
    
    print("\n" + "="*70)
    print("✅ 数据处理完成！")
    print("="*70)


if __name__ == "__main__":
    main()
